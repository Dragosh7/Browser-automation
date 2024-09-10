import os
from pathlib import Path

import openpyxl
import random
import time
import requests
from robocorp import browser
from robocorp.tasks import task
from RPA.Excel.Files import Files as Excel

FILE_NAME = "challenge.xlsx"
EXCEL_URL = f"https://rpachallenge.com/assets/downloadFiles/{FILE_NAME}"
OUTPUT_DIR = Path(os.getenv("ROBOT_ARTIFACTS", "output"))


@task
def solve_challenge():
    """
    Main task which solves the RPA challenge!

    Downloads the source data Excel file and uses Playwright to fill the entries inside
    rpachallenge.com.
    """
    browser.configure(
        browser_engine="chromium", 
        screenshot="only-on-failure", 
        headless=True 
    )
    try:
        # Reads a table from an Excel file hosted online.
        excel_file = download_file(
            EXCEL_URL, target_dir=OUTPUT_DIR, target_filename=FILE_NAME
        )
        excel = Excel()
        excel.open_workbook(excel_file)
        rows = excel.read_worksheet_as_table("Sheet1", header=True)

        # Surf the automation challenge website and fill in information from the table
        #  extracted above.
        page = browser.goto("https://rpachallenge.com/")
        page.click("button:text('Start')")
        for row in rows:
            fill_and_submit_form(row, page=page)
        element = page.locator("css=div.congratulations")
        browser.screenshot(element)
    finally:
        # A place for teardown and cleanups. (Playwright handles browser closing)
        print("Automation finished!")


def download_file(url: str, *, target_dir: Path, target_filename: str) -> Path:
    """
    Downloads a file from the given URL into a custom folder & name.

    Args:
        url: The target URL from which we'll download the file.
        target_dir: The destination directory in which we'll place the file.
        target_filename: The local file name inside which the content gets saved.

    Returns:
        Path: A Path object pointing to the downloaded file.
    """
    # Obtain the content of the file hosted online.
    response = requests.get(url)
    response.raise_for_status()  # this will raise an exception if the request fails
    # Write the content of the request response to the target file.
    target_dir.mkdir(exist_ok=True)
    local_file = target_dir / target_filename
    local_file.write_bytes(response.content)
    return local_file


def fill_and_submit_form(row: dict, *, page: browser.Page):
    """
    Fills a single form with the information of a single row from the table.

    Args:
        row: One row from the generated table out of the input Excel file.
        page: The page object over which the browser interactions are done.
    """
    field_data_map = {
        "labelFirstName": "First Name",
        "labelLastName": "Last Name",
        "labelCompanyName": "Company Name",
        "labelRole": "Role in Company",
        "labelAddress": "Address",
        "labelEmail": "Email",
        "labelPhone": "Phone Number",
    }
    for field, key in field_data_map.items():
        page.fill(f"//input[@ng-reflect-name='{field}']", str(row[key]))
    page.click("input:text('Submit')")

@task
def back_in_stock_notification():
    browser.configure(
        browser_engine="chromium", 
        screenshot="only-on-failure", 
        headless=False 
    )

    page = browser.goto("https://www.ebay.com/")

    try:
        page.click('button:has-text("Accept all")')
    except:
        pass 

    page.wait_for_timeout(2000)

    search_term1 = "bmw"
    search_term2 = "m7658"
    while (1):
        page.fill("input[name='_nkw']", search_term1 + " " + search_term2)
        page.click("input[type='submit']")


        page.wait_for_timeout(2000)  # Wait for images to load
        page.wait_for_selector("#srp-river-results")
        results = page.locator("#srp-river-results .s-item")

        product_found = False
        for i in range(results.count()):
            title = results.nth(i).locator(".s-item__title").inner_text().lower()

            if search_term1 in title and search_term2 in title:
                print("available")
                product_found=True
                break

        if product_found:
            break
        else:
            print("No matching results found.")

            wait_time = random.randint(1800, 3600)  # Random between 30 to 60 minutes
            print(f"Waiting for {wait_time/60} minutes before rechecking...")
            time.sleep(wait_time)
            page.goto("https://www.ebay.com/")

    
def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return workbook, sheet

def check_multiple_stock_notification(sheet, workbook, file_path):
    browser.configure(
        browser_engine="chromium", 
        screenshot="only-on-failure", 
        headless=False
    )

    #page = browser.goto("https://altex.ro/telefon-apple-iphone-14-pro-5g-128gb-space-black/cpd/SMTIP14PR1BK/")
    

    for row in range(2, sheet.max_row + 1):
        product_url = sheet.cell(row, 1).value  # Get the product URL from the first column
        if not product_url:
            continue
        print(product_url)
        
        page = browser.goto(product_url)
        page.wait_for_timeout(2000) 


        # Check stock
        try:
            print("Waiting for stock element...")
            # if(page.wait_for_selector('//div[contains(text(), "Produsul nu mai face parte din catalogul magazinului.")]', timeout=5000)):
            #   continue
            page.wait_for_selector('//div[contains(text(), "in stoc")]', timeout=5000)
            stock_element = page.query_selector('//div[contains(text(), "in stoc")]')
            if stock_element:
                stock_text = stock_element.inner_text().lower()
                print(f"Stock element found: {stock_text}")
                if "in stoc" in stock_text:
                    page.wait_for_selector('div.leading-none span span.Price-int', timeout=5000)
                    price_element = page.query_selector('div.leading-none span span.Price-int')
                    if price_element:
                        stock_price = price_element.inner_text()
                        print(f"Stock price found: {stock_price}")
                    else:
                        stock_price = "-"
                        print("Price element not found.")
                else:
                    stock_price = "-"
                    print("Stock status not 'in stoc'.")
            else:
                stock_price = "-"
                print("Stock element not found.")

            sheet.cell(row, 2).value = stock_price + ' Lei'  # Update the stock price in Excel
        except Exception as e:
            print(f"Error checking stock: {e}")
            sheet.cell(row, 2).value = "-"  # not available  (exceptions case)

        # Check resealed stock
        try:
            print("Waiting for resealed button...")
            page.wait_for_selector('a[href="#resigilate"]', timeout=5000)
            resealed_button = page.query_selector('a[href="#resigilate"]')
            if resealed_button:
                print("Resealed button found, clicking...")
                resealed_button.click()
                page.wait_for_timeout(1000)  # Adjust timeout as needed

                first_li_selector = 'ul > li:first-of-type'
                price_selector_js_path = '#resigilate > ul > li:nth-child(1) > div.lg\\:w-5\\/12.lg\\:pl-8.flex.flex-row.items-center.space-x-5.sm\\:space-x-0.justify-between > div.font-bold.inline-block.text-28px > div.leading-none.text-red-brand.text-\\[\\#39ab4a\\].-tracking-0\\.48.lg\\:-tracking-0\\.56 > span > span.Price-int.leading-none'

                page.wait_for_selector(price_selector_js_path, timeout=5000)
                price_element = page.query_selector(price_selector_js_path)
                if price_element:
                    price = price_element.inner_text()
                    print(f"Resealed price found: {price}")
                else:
                    price = "-"
                    print("Resealed price element not found.")
            
                sheet.cell(row, 3).value = price + ' Lei'  # Update the stock price in Excel
            else:
                sheet.cell(row, 3).value = "-"
                print("Resealed button not found.")
        except Exception as e:
            print(f"Error checking resealed stock: {e}")
            sheet.cell(row, 3).value = "-"  # not available  (exceptions case)

        # Save and close
        workbook.save(file_path)
        print(f"Updated Excel file for URL: {product_url}")

    page.close()
    print("Finished processing all URLs.")

@task
def task_3a_advanced():
    file_path = "C:/Users/z004zy5d/Desktop/products_data.xlsx" 
    workbook, sheet = read_excel(file_path)
    check_multiple_stock_notification(sheet,workbook,file_path)